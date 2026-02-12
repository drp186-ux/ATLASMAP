#!/usr/bin/env python3
import json
import re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "data" / "partners.xlsx"
OUT_ROUTES = ROOT / "data" / "routes.json"
OUT_LOCS = ROOT / "data" / "locations.json"

PALETTE = [
  "#1f77b4","#ff7f0e","#2ca02c","#d62728","#9467bd",
  "#8c564b","#e377c2","#7f7f7f","#bcbd22","#17becf",
  "#393b79","#637939","#8c6d31","#843c39","#7b4173"
]

def split_points(s: str):
    s = str(s).strip()
    s = s.replace("->", "→").replace("—", "→").replace(" – ", "→").replace(" — ", "→")
    s = re.sub(r"\s-\s", "→", s)
    return [x.strip() for x in s.split("→") if x.strip()]

def split_routes(raw: str):
    raw = str(raw).strip()
    parts = re.split(r"[\n;|]+", raw)
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        sub = re.split(r"\s{2,}(?=\S+\s*→)", p)
        for sp in sub:
            pts = split_points(sp)
            if len(pts) >= 2:
                out.append(pts)
    return out

def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        prev = None
        for r in range(2, ws.max_row + 1):
            car = ws.cell(r, 1).value
            routes = ws.cell(r, 2).value
            if car is None and routes and prev:
                car = prev
            if car and routes:
                rows.append((str(car).strip(), str(routes).strip()))
                prev = str(car).strip()
            elif car:
                prev = str(car).strip()

    seen = set()
    routes = []
    for car, raw in rows:
        for pts in split_routes(raw):
            key = (car, tuple(pts))
            if key in seen:
                continue
            seen.add(key)
            routes.append((car, pts))

    # Fixes requested earlier
    final_routes = []
    for car, pts in routes:
        new = []
        for p in pts:
            pc = p.strip()
            if pc.lower() in ["котла", "котлп"]:
                new.append("Котлас")
                continue
            if car.strip().upper() == "СКС":
                if pc.lower() == "полоцк":
                    new.append("Полоцк, Беларусь"); continue
                if pc.lower() == "минск":
                    new.append("Минск, Беларусь"); continue
            new.append(pc)
        final_routes.append((car, new))

    carriers_sorted = sorted({c for c, _ in final_routes})
    color_map = {c: PALETTE[i % len(PALETTE)] for i, c in enumerate(carriers_sorted)}

    routes_json = []
    for i, (car, pts) in enumerate(final_routes, start=1):
        routes_json.append({
            "route_id": f"RTE_{i:04d}",
            "carrier": car,
            "color": color_map[car],
            "points": pts,
            "route_name": " → ".join(pts),
        })

    loc_names = sorted({p for _, pts in final_routes for p in pts})
    locations_json = [{"name": n, "country": ("BY" if "Беларусь" in n else "RU"), "lat": None, "lon": None} for n in loc_names]

    OUT_ROUTES.write_text(json.dumps(routes_json, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_LOCS.write_text(json.dumps(locations_json, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK: {len(carriers_sorted)} carriers, {len(routes_json)} routes, {len(loc_names)} locations")

if __name__ == "__main__":
    main()
